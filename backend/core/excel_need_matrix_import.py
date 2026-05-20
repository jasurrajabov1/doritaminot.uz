# coding: utf-8
import re
from datetime import date
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .models import Drug, Institution, MonthlyIssue, NeedAddition, NeedRow


CONTROL_GROUP_LABELS = {
    "general": "Оддий препарат",
    "narcotic": "Гиёҳвандлик воситалари",
    "psychotropic": "Психотроп моддалар",
    "strong": "Кучли таъсир қилувчи",
    "precursor": "Прекурсорлар",
}


def _clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def _norm(value):
    text = _clean_text(value).lower()
    text = text.replace("ў", "у").replace("қ", "к").replace("ғ", "г").replace("ҳ", "х")
    text = re.sub(r"\s+", " ", text)
    return text


def _to_int(value, default=None):
    text = _clean_text(value)
    if not text:
        return default
    try:
        return int(float(text.replace(",", ".")))
    except Exception:
        return default


def _to_bool(value):
    text = _norm(value)
    return text in {"1", "true", "yes", "ha", "ҳа", "xa", "on"}


def _col_to_index(value, default=None):
    text = _clean_text(value)
    if not text:
        return default
    if text.isdigit():
        return int(text)
    try:
        return column_index_from_string(text.upper())
    except Exception:
        return default


def _decimal_or_none(value):
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        return Decimal(str(value))

    text = _clean_text(value)
    if not text or text in {"—", "-", "–"}:
        return None

    text = text.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)

    if not text or text in {"-", "."}:
        return None

    try:
        return Decimal(text).quantize(Decimal("0.001"))
    except (InvalidOperation, ValueError):
        return None


def _decimal_zero(value):
    parsed = _decimal_or_none(value)
    return parsed if parsed is not None else Decimal("0.000")


def _make_merged_value_map(ws):
    result = {}

    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                result[(row, col)] = value

    return result


def _cell_value(ws, merged_values, row, col):
    value = ws.cell(row=row, column=col).value
    if value is not None:
        return value
    return merged_values.get((row, col))


def _cell_text(ws, merged_values, row, col):
    return _clean_text(_cell_value(ws, merged_values, row, col))


def _control_group_from_sheet(sheet_name):
    text = _norm(sheet_name)

    if "гиёх" in text or "гиёхванд" in text or "narc" in text:
        return "narcotic"

    if "псих" in text or "psych" in text:
        return "psychotropic"

    if "кучли" in text or "таъсир" in text or "strong" in text:
        return "strong"

    if "прекурс" in text or "precursor" in text:
        return "precursor"

    return "general"


def _guess_mnn_name(drug_title):
    text = _clean_text(drug_title)
    text = re.sub(r"\([^)]*\)", " ", text)

    # дозалар ва бирликларни тахминий олиб ташлаш
    text = re.sub(r"\b\d+[,.]?\d*\s*%?\b", " ", text)
    text = re.sub(
        r"\b(мг|гр|г|мл|л|мкг|mcg|mg|g|ml|tab|tabs|табл|таб|амп|ампула|фл|капс|капсула|сироп|эритма|раствор|тюб|крем|мазь|гель|порошок|спрей|томчи)\b",
        " ",
        text,
        flags=re.IGNORECASE,
    )

    text = re.sub(r"[/,;]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" -–—")

    if not text:
        return _clean_text(drug_title)

    parts = text.split()
    if len(parts) > 3:
        return " ".join(parts[:3])

    return text


def _extract_title_and_unit(raw_title, fallback_unit=""):
    text = _clean_text(raw_title)
    unit = ""

    match = re.search(r"\(([^)]*)\)\s*$", text)
    if match:
        unit = _clean_text(match.group(1))
        text = re.sub(r"\([^)]*\)\s*$", "", text).strip()

    return text, unit or _clean_text(fallback_unit)


_SKIP_TITLE_PARTS = {
    "инн",
    "улч",
    "улч.",
    "бир",
    "эҳтиёж",
    "эхтиёж",
    "берилган",
    "жами",
    "йиллик",
    "кушимча",
    "кушимча",
    "дпм",
    "амб",
    "рец",
    "даволаш",
    "профилактика",
    "муассаса",
    "номланиши",
    "1-жадвал",
    "жадвал",
    "т/р",
    "№",
}


def _candidate_drug_title(ws, merged_values, col):
    for row in (2, 3, 1):
        text = _cell_text(ws, merged_values, row, col)
        if not text:
            continue

        normalized = _norm(text)

        if any(part in normalized for part in _SKIP_TITLE_PARTS):
            continue

        if len(text) < 3:
            continue

        return text

    return ""


def _header_text_for_col(ws, merged_values, col):
    values = []
    for row in range(4, 7):
        text = _cell_text(ws, merged_values, row, col)
        if text:
            values.append(text)
    return _norm(" ".join(values))


def _detect_drug_blocks(ws, merged_values, start_col=5):
    groups = []
    current = None

    for col in range(start_col, ws.max_column + 1):
        title = _candidate_drug_title(ws, merged_values, col)

        if not title:
            continue

        if current and _norm(current["raw_title"]) == _norm(title):
            current["end_col"] = col
        else:
            current = {
                "raw_title": title,
                "start_col": col,
                "end_col": col,
            }
            groups.append(current)

    detected = []

    for group in groups:
        base_col = None
        add_col = None
        total_col = None
        issued_dpm_col = None
        issued_amb_col = None

        for col in range(group["start_col"], group["end_col"] + 1):
            header = _header_text_for_col(ws, merged_values, col)

            if not header:
                continue

            is_need = "эхтиёж" in header or "эҳтиёж" in header
            is_issued = "берилган" in header

            if "йиллик" in header and is_need:
                base_col = col
            elif ("кушимча" in header or "қушимча" in header or "қўшимча" in header) and is_need:
                add_col = col
            elif "жами" in header and is_need:
                total_col = col
            elif "дпм" in header and is_issued:
                issued_dpm_col = col
            elif ("амб" in header or "рец" in header) and is_issued:
                issued_amb_col = col

        if not any([base_col, add_col, total_col, issued_dpm_col, issued_amb_col]):
            continue

        title, unit = _extract_title_and_unit(group["raw_title"])

        if not title:
            continue

        detected.append({
            "drug_title": title,
            "unit": unit,
            "start_col": group["start_col"],
            "end_col": group["end_col"],
            "base_col": base_col,
            "add_col": add_col,
            "total_col": total_col,
            "issued_dpm_col": issued_dpm_col,
            "issued_amb_col": issued_amb_col,
        })

    return detected


def _sheet_names_from_mapping(wb, mapping):
    raw = _clean_text(mapping.get("sheet_name") or mapping.get("sheet_names") or "")

    if not raw:
        return list(wb.sheetnames)

    requested = [x.strip() for x in re.split(r"[;,]", raw) if x.strip()]
    existing = []

    for item in requested:
        if item in wb.sheetnames:
            existing.append(item)
            continue

        lower = _norm(item)
        match = next((name for name in wb.sheetnames if _norm(name) == lower), None)
        if match:
            existing.append(match)

    return existing


def _read_rows(uploaded_file, mapping):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True)

    sheet_names = _sheet_names_from_mapping(wb, mapping)
    if not sheet_names:
        return [], {
            "sheets": wb.sheetnames,
            "detail": "Excel варағи топилмади.",
        }

    start_row = _to_int(mapping.get("start_row"), 7)
    end_row_raw = _to_int(mapping.get("end_row"), None)

    inst_name_col = _col_to_index(mapping.get("institution_name_col"), 2)
    inst_inn_col = _col_to_index(mapping.get("institution_inn_col"), 3)
    unit_col = _col_to_index(mapping.get("unit_col"), 4)

    rows = []
    meta = {
        "sheets": wb.sheetnames,
        "selected_sheets": sheet_names,
        "blocks": [],
    }

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        merged_values = _make_merged_value_map(ws)

        end_row = end_row_raw or ws.max_row
        control_group = _control_group_from_sheet(sheet_name)
        blocks = _detect_drug_blocks(ws, merged_values, start_col=5)

        meta["blocks"].append({
            "sheet_name": sheet_name,
            "control_group": control_group,
            "control_group_label": CONTROL_GROUP_LABELS.get(control_group, control_group),
            "drug_count": len(blocks),
            "drugs": [
                {
                    "drug_title": block["drug_title"],
                    "start_col": block["start_col"],
                    "end_col": block["end_col"],
                }
                for block in blocks
            ][:80],
        })

        for row_number in range(start_row, end_row + 1):
            institution_name = _cell_text(ws, merged_values, row_number, inst_name_col)
            institution_inn = re.sub(r"\D", "", _cell_text(ws, merged_values, row_number, inst_inn_col))
            row_unit = _cell_text(ws, merged_values, row_number, unit_col)

            if not institution_name and not institution_inn:
                continue

            for block in blocks:
                base_value = _decimal_or_none(
                    _cell_value(ws, merged_values, row_number, block.get("base_col"))
                ) if block.get("base_col") else None

                add_value = _decimal_or_none(
                    _cell_value(ws, merged_values, row_number, block.get("add_col"))
                ) if block.get("add_col") else None

                total_value = _decimal_or_none(
                    _cell_value(ws, merged_values, row_number, block.get("total_col"))
                ) if block.get("total_col") else None

                issued_dpm = _decimal_or_none(
                    _cell_value(ws, merged_values, row_number, block.get("issued_dpm_col"))
                ) if block.get("issued_dpm_col") else None

                issued_amb = _decimal_or_none(
                    _cell_value(ws, merged_values, row_number, block.get("issued_amb_col"))
                ) if block.get("issued_amb_col") else None

                base_need = base_value
                additional_need = add_value or Decimal("0.000")

                if base_need is None:
                    if total_value is not None:
                        base_need = total_value - additional_need
                        if base_need < 0:
                            base_need = Decimal("0.000")
                    else:
                        base_need = Decimal("0.000")

                issued_qty = (issued_dpm or Decimal("0.000")) + (issued_amb or Decimal("0.000"))

                if (
                    base_need <= 0
                    and additional_need <= 0
                    and issued_qty <= 0
                    and (total_value is None or total_value <= 0)
                ):
                    continue

                rows.append({
                    "sheet_name": sheet_name,
                    "row_number": row_number,
                    "institution_name": institution_name,
                    "institution_inn": institution_inn,
                    "drug_title": block["drug_title"],
                    "unit": block.get("unit") or row_unit,
                    "control_group": control_group,
                    "base_need": base_need.quantize(Decimal("0.001")),
                    "additional_need": additional_need.quantize(Decimal("0.001")),
                    "total_excel": total_value.quantize(Decimal("0.001")) if total_value is not None else None,
                    "issued_qty": issued_qty.quantize(Decimal("0.001")),
                })

    return rows, meta


def _find_institution(data):
    inn = _clean_text(data.get("institution_inn"))
    name = _clean_text(data.get("institution_name"))

    if inn:
        found = Institution.objects.filter(inn=inn).first()
        if found:
            return found

    if name:
        found = Institution.objects.filter(name__iexact=name).first()
        if found:
            return found

        compact = re.sub(r"\s+", " ", name).strip()
        found = Institution.objects.filter(name__icontains=compact[:40]).first()
        if found:
            return found

    return None


def _find_drug(title):
    title = _clean_text(title)

    if not title:
        return None

    if hasattr(Drug, "full_name"):
        found = Drug.objects.filter(full_name__iexact=title).first()
        if found:
            return found

    return Drug.objects.filter(name__iexact=title).first()


def _create_or_update_drug(data, update_existing, commit):
    title = _clean_text(data["drug_title"])
    unit = _clean_text(data.get("unit"))
    control_group = data.get("control_group") or "general"

    drug = _find_drug(title)

    if drug:
        action = "drug_exists"

        if update_existing:
            changed = False

            if unit and hasattr(drug, "unit") and not (drug.unit or "").strip():
                drug.unit = unit
                changed = True

            if hasattr(drug, "mnn_name") and not (drug.mnn_name or "").strip():
                drug.mnn_name = _guess_mnn_name(title)
                changed = True

            if hasattr(drug, "control_group"):
                current = getattr(drug, "control_group", "") or "general"
                if current == "general" and control_group != "general":
                    setattr(drug, "control_group", control_group)
                    changed = True

            if changed:
                action = "drug_update"
                if commit:
                    drug.save()

        return drug, action

    action = "drug_create"

    if not commit:
        return None, action

    payload = {
        "name": title,
        "is_active": True,
    }

    if hasattr(Drug, "mnn_name"):
        payload["mnn_name"] = _guess_mnn_name(title)

    if hasattr(Drug, "unit"):
        payload["unit"] = unit

    if hasattr(Drug, "control_group"):
        payload["control_group"] = control_group

    return Drug.objects.create(**payload), action


def _row_result(data, ok, action, message="", errors=""):
    return {
        "sheet_name": data.get("sheet_name"),
        "row_number": data.get("row_number"),
        "ok": bool(ok),
        "action": action,
        "message": message,
        "errors": errors,
        "data": {
            "institution_name": data.get("institution_name"),
            "institution_inn": data.get("institution_inn"),
            "drug_title": data.get("drug_title"),
            "unit": data.get("unit"),
            "control_group": data.get("control_group"),
            "control_group_label": CONTROL_GROUP_LABELS.get(data.get("control_group"), data.get("control_group")),
            "base_need": float(data.get("base_need") or 0),
            "additional_need": float(data.get("additional_need") or 0),
            "total_excel": float(data.get("total_excel")) if data.get("total_excel") is not None else None,
            "issued_qty": float(data.get("issued_qty") or 0),
        },
    }


def _validate_row(data, update_existing=True):
    messages = []
    errors = []

    institution = _find_institution(data)
    if not institution:
        errors.append(
            f"Муассаса топилмади. Excel номи: {data.get('institution_name')}; ИНН: {data.get('institution_inn') or '—'}"
        )

    total_calc = (data["base_need"] or Decimal("0")) + (data["additional_need"] or Decimal("0"))

    if data.get("total_excel") is not None:
        diff = abs(total_calc - data["total_excel"])
        if diff > Decimal("0.010"):
            messages.append(
                f"Excel жами эҳтиёж билан ҳисобланган жамида фарқ бор. Excel: {data['total_excel']}, ҳисоб: {total_calc}"
            )

    if data["issued_qty"] > total_calc:
        errors.append(
            f"Берилган миқдор жами эҳтиёждан ошиб кетган. Жами эҳтиёж: {total_calc}, берилган: {data['issued_qty']}"
        )

    drug = _find_drug(data["drug_title"])
    drug_action = "drug_exists" if drug else "drug_create"

    if institution and drug:
        need_exists = NeedRow.objects.filter(
            institution=institution,
            drug=drug,
            year=int(data.get("year") or date.today().year),
        ).exists()
    else:
        need_exists = False

    need_action = "need_update" if need_exists and update_existing else "need_create"

    action = "; ".join([drug_action, need_action])

    if errors:
        return _row_result(data, False, action, " | ".join(messages), " | ".join(errors))

    return _row_result(data, True, action, " | ".join(messages), "")


def _save_row(data, update_existing=True, user=None):
    validation = _validate_row(data, update_existing=update_existing)
    if not validation["ok"]:
        return validation

    year = int(data.get("year") or date.today().year)
    institution = _find_institution(data)
    drug, drug_action = _create_or_update_drug(data, update_existing=update_existing, commit=True)

    need_row, need_created = NeedRow.objects.get_or_create(
        institution=institution,
        drug=drug,
        year=year,
        defaults={
            "dpm_need": data["base_need"],
            "amb_rec_need": Decimal("0.000"),
        },
    )

    need_action = "need_create" if need_created else "need_exists"

    if not need_created and update_existing:
        if need_row.dpm_need != data["base_need"] or need_row.amb_rec_need != Decimal("0.000"):
            need_row.dpm_need = data["base_need"]
            need_row.amb_rec_need = Decimal("0.000")
            need_row.save()
            need_action = "need_update"

    addition_action = "addition_skip"
    doc_number = f"Excel import: {data.get('sheet_name')}"[:120]
    existing_addition = NeedAddition.objects.filter(
        need_row=need_row,
        doc_number=doc_number,
    ).first()

    if data["additional_need"] > 0:
        reason = getattr(NeedAddition, "REASON_SSV_ORDER", "ssv_order")

        if existing_addition:
            existing_addition.dpm_need_add = data["additional_need"]
            existing_addition.amb_rec_need_add = Decimal("0.000")
            existing_addition.addition_date = date.today()
            existing_addition.reason = reason
            existing_addition.comment = "Excel матрицадан импорт қилинган қўшимча эҳтиёж."
            existing_addition.is_active = True
            existing_addition.cancel_reason = ""
            existing_addition.save()
            addition_action = "addition_update"
        else:
            NeedAddition.objects.create(
                need_row=need_row,
                institution=institution,
                drug=drug,
                year=year,
                dpm_need_add=data["additional_need"],
                amb_rec_need_add=Decimal("0.000"),
                addition_date=date.today(),
                reason=reason,
                doc_number=doc_number,
                comment="Excel матрицадан импорт қилинган қўшимча эҳтиёж.",
                is_active=True,
                created_by=user if user and user.is_authenticated else None,
            )
            addition_action = "addition_create"
    elif existing_addition and update_existing and existing_addition.is_active:
        existing_addition.is_active = False
        existing_addition.cancel_reason = "Excel импортда қўшимча эҳтиёж 0 бўлгани учун бекор қилинди."
        existing_addition.save()
        addition_action = "addition_cancel"

    issue_action = "issued_skip"
    existing_issue = MonthlyIssue.objects.filter(
        institution=institution,
        drug=drug,
        year=year,
    ).first()

    if existing_issue:
        if update_existing:
            existing_issue.issued_qty = data["issued_qty"]
            existing_issue.save(update_fields=["issued_qty"])
            issue_action = "issued_update"
        else:
            issue_action = "issued_exists_skip"
    else:
        if data["issued_qty"] > 0:
            MonthlyIssue.objects.create(
                institution=institution,
                drug=drug,
                year=year,
                issued_qty=data["issued_qty"],
            )
            issue_action = "issued_create"

    action = "; ".join([drug_action, need_action, addition_action, issue_action])
    return _row_result(data, True, action, "", "")



def _summary(rows, meta, commit=False, detail=""):
    rows = rows or []
    total = len(rows)

    ok_rows = [row for row in rows if row.get("ok")]
    error_rows = [row for row in rows if not row.get("ok")]

    ok_count = len(ok_rows)
    error_count = len(error_rows)

    created = sum(
        1
        for row in ok_rows
        if "create" in (row.get("action") or "")
    )

    updated = sum(
        1
        for row in ok_rows
        if "update" in (row.get("action") or "")
    )

    skipped = error_count + sum(
        1
        for row in ok_rows
        if "skip" in (row.get("action") or "")
    )

    if detail:
        resolved_detail = detail
    elif commit:
        resolved_detail = (
            f"Базага {ok_count} та тўғри қатор импорт қилинди. "
            f"{error_count} та хатоли қатор ўтказиб юборилди."
        )
    else:
        resolved_detail = (
            f"Excel текширилди. Тўғри: {ok_count}. Хато: {error_count}."
        )

    return {
        # ok=True энди “камида битта сақлаш мумкин бўлган қатор бор” дегани.
        # Ҳамма хатосиз дегани эмас. Ҳамма хатосизлик учун has_errors=False бўлади.
        "ok": ok_count > 0,
        "has_errors": error_count > 0,
        "can_commit": ok_count > 0,
        "commit": bool(commit),
        "detail": resolved_detail,

        "total": total,
        "ok_count": ok_count,
        "errors_count": error_count,
        "created": created,
        "updated": updated,
        "skipped": skipped,

        "meta": meta,
        "rows": rows,
        "ok_rows": ok_rows,
        "error_rows": error_rows,
    }


def process_need_matrix_import(uploaded_file, mapping, commit=False, update_existing=True, user=None):
    mapping = dict(mapping or {})
    year = _to_int(mapping.get("year"), date.today().year)
    mapping["year"] = year

    try:
        data_rows, meta = _read_rows(uploaded_file, mapping)
    except Exception as exc:
        return {
            "ok": False,
            "has_errors": True,
            "can_commit": False,
            "commit": False,
            "detail": f"Excel ўқишда хато: {exc}",
            "total": 0,
            "ok_count": 0,
            "errors_count": 1,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "meta": {},
            "rows": [],
            "ok_rows": [],
            "error_rows": [],
        }

    for row in data_rows:
        row["year"] = year

    if not data_rows:
        return {
            "ok": False,
            "has_errors": True,
            "can_commit": False,
            "commit": False,
            "detail": "Импорт қилинадиган қатор топилмади. Бошланиш қатори, варақ номи ёки Excel сарлавҳаларини текширинг.",
            "total": 0,
            "ok_count": 0,
            "errors_count": 1,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "meta": meta,
            "rows": [],
            "ok_rows": [],
            "error_rows": [],
        }

    preview_rows = [
        _validate_row(row, update_existing=update_existing)
        for row in data_rows
    ]

    preview_summary = _summary(preview_rows, meta, commit=False)

    if not commit:
        return preview_summary

    valid_pairs = [
        (data_row, preview_row)
        for data_row, preview_row in zip(data_rows, preview_rows)
        if preview_row.get("ok")
    ]

    skipped_error_rows = [
        preview_row
        for preview_row in preview_rows
        if not preview_row.get("ok")
    ]

    if not valid_pairs:
        preview_summary["detail"] = "Ҳамма қаторда хато бор. Базага импорт қилинмади."
        preview_summary["ok"] = False
        preview_summary["can_commit"] = False
        preview_summary["commit"] = False
        return preview_summary

    saved_rows = []

    # Муҳим: ҳар бир тўғри қатор алоҳида transaction билан сақланади.
    # Бир қаторда кейинчалик DB хатоси чиқса, қолган OK қаторлар бекор қилинмайди.
    for data_row, preview_row in valid_pairs:
        try:
            with transaction.atomic():
                saved = _save_row(
                    data_row,
                    update_existing=update_existing,
                    user=user,
                )

            saved_rows.append(saved)

        except Exception as exc:
            failed = dict(preview_row)
            failed["ok"] = False
            failed["action"] = f"{failed.get('action') or 'save'}; skip"
            failed["message"] = "Бу қатор сақланмади."
            failed["errors"] = str(exc)
            saved_rows.append(failed)

    final_rows = saved_rows + skipped_error_rows
    saved_ok = sum(1 for row in saved_rows if row.get("ok"))
    final_errors = sum(1 for row in final_rows if not row.get("ok"))

    detail = (
        f"Базага {saved_ok} та тўғри қатор импорт қилинди. "
        f"{final_errors} та хатоли қатор ўтказиб юборилди."
    )

    return _summary(final_rows, meta, commit=True, detail=detail)

