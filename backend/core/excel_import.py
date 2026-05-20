# coding: utf-8
import re
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from django.db import transaction

from .models import AuditLog, Institution
from .serializers import InstitutionSerializer


IMPORT_TYPES = {
    "institutions": {
        "label": "Муассасалар",
        "page_code": "institutions",
    },
}


HEADER_ALIASES = {
    "name": {
        "name",
        "nomi",
        "муассаса",
        "муассаса номи",
        "муассаса номланиши",
        "даволаш-профилактика муассасаларининг номланиши",
        "institution",
        "institution name",
    },
    "inn": {
        "inn",
        "инн",
        "stir",
        "tin",
        "солиқ",
    },
    "address": {
        "address",
        "манзил",
        "адрес",
    },
    "is_active": {
        "active",
        "фаол",
        "is_active",
    },
}


def _cell_text(value):
    if value is None:
        return ""

    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    return str(value).strip()


def _normalize_header(value):
    text = _cell_text(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_inn(value):
    text = _cell_text(value).strip()

    if not text:
        return ""

    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    digits = re.sub(r"\D", "", text)

    if len(digits) == 9:
        return digits

    return text


def _truthy(value):
    text = _cell_text(value).strip().lower()
    if not text:
        return True
    return text in {"1", "true", "yes", "ҳа", "ха", "active", "фаол"}


def _col_to_index(value):
    text = _cell_text(value).strip()

    if not text:
        return None

    if text.isdigit():
        return int(text)

    return column_index_from_string(text.upper())


def _get_cell(ws, row_number, col_value):
    col_index = _col_to_index(col_value)
    if not col_index:
        return ""
    return _cell_text(ws.cell(row=row_number, column=col_index).value)


def _detect_headers(ws):
    max_scan_rows = min(ws.max_row or 1, 20)
    max_scan_cols = min(ws.max_column or 1, 80)

    for row_number in range(1, max_scan_rows + 1):
        found = {}
        for col_number in range(1, max_scan_cols + 1):
            header = _normalize_header(ws.cell(row=row_number, column=col_number).value)
            if not header:
                continue

            for field, aliases in HEADER_ALIASES.items():
                if header in aliases and field not in found:
                    found[field] = col_number

        if "name" in found:
            return row_number, found

    return None, {}


def _row_result(row_number, ok, action, message, data=None, errors=None, object_id=None):
    return {
        "row_number": row_number,
        "ok": ok,
        "action": action,
        "message": message,
        "data": data or {},
        "errors": errors or "",
        "object_id": object_id,
    }


def _summarize(import_type, rows, commit=False, rolled_back=False):
    valid_rows = [r for r in rows if r.get("ok")]
    error_rows = [r for r in rows if not r.get("ok")]

    return {
        "ok": len(error_rows) == 0,
        "import_type": import_type,
        "commit": bool(commit),
        "rolled_back": bool(rolled_back),
        "total": len(rows),
        "valid": len(valid_rows),
        "errors_count": len(error_rows),
        "created": sum(1 for r in valid_rows if r.get("action") == "create"),
        "updated": sum(1 for r in valid_rows if r.get("action") == "update"),
        "skipped": sum(1 for r in rows if r.get("action") == "skip"),
        "rows": rows[:500],
    }


def _validate_or_save_institution(row_number, data, update_existing=False, commit=False, user=None):
    name = (data.get("name") or "").strip()
    inn = _normalize_inn(data.get("inn"))
    address = (data.get("address") or "").strip()
    is_active = _truthy(data.get("is_active"))

    payload = {
        "name": name,
        "inn": inn,
        "address": address,
        "is_active": is_active,
    }

    if not name and not inn:
        return _row_result(row_number, True, "skip", "Бўш қатор ўтказилди.", payload)

    if not name:
        return _row_result(row_number, False, "error", "Муассаса номи бўш.", payload, "Муассаса номи мажбурий.")

    existing = None
    if inn:
        existing = Institution.objects.filter(inn=inn).first()

    if not existing:
        existing = Institution.objects.filter(name__iexact=name).first()

    if existing and not update_existing:
        return _row_result(
            row_number,
            False,
            "duplicate",
            "Бу муассаса базада мавжуд.",
            payload,
            "Мавжуд ёзувни янгилаш учун checkbox'ни белгиланг.",
            existing.id,
        )

    serializer = InstitutionSerializer(instance=existing, data=payload, partial=bool(existing))
    if not serializer.is_valid():
        return _row_result(
            row_number,
            False,
            "error",
            "Validation хато.",
            payload,
            serializer.errors,
            existing.id if existing else None,
        )

    action = "update" if existing else "create"

    if commit:
        item = serializer.save()

        AuditLog.objects.create(
            actor=user if user and user.is_authenticated else None,
            action=action,
            target_type="Муассаса",
            target_id=str(item.id),
            target_repr=str(item),
            description="Excel import орқали муассаса қўшилди." if action == "create" else "Excel import орқали муассаса янгиланди.",
            extra_data={
                "source": "excel_import",
                "row_number": row_number,
                "name": item.name,
                "inn": item.inn,
                "address": item.address,
                "is_active": item.is_active,
            },
        )

        return _row_result(row_number, True, action, "OK", payload, object_id=item.id)

    return _row_result(row_number, True, action, "OK", payload, object_id=existing.id if existing else None)


def _read_institution_rows_template(ws):
    header_row, headers = _detect_headers(ws)

    if not header_row or "name" not in headers:
        raise ValueError("Биринчи қаторларда устун номлари топилмади. Manual mapping режимидан фойдаланинг.")

    rows = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        rows.append({
            "row_number": row_number,
            "name": _cell_text(ws.cell(row=row_number, column=headers.get("name")).value) if headers.get("name") else "",
            "inn": _cell_text(ws.cell(row=row_number, column=headers.get("inn")).value) if headers.get("inn") else "",
            "address": _cell_text(ws.cell(row=row_number, column=headers.get("address")).value) if headers.get("address") else "",
            "is_active": _cell_text(ws.cell(row=row_number, column=headers.get("is_active")).value) if headers.get("is_active") else "ҳа",
        })
    return rows


def _read_institution_rows_manual(ws, mapping):
    start_row = int(mapping.get("start_row") or 1)
    end_row = mapping.get("end_row") or ""
    end_row = int(end_row) if str(end_row).strip() else ws.max_row

    name_col = mapping.get("name_col") or mapping.get("name") or ""
    inn_col = mapping.get("inn_col") or mapping.get("inn") or ""
    address_col = mapping.get("address_col") or mapping.get("address") or ""
    active_col = mapping.get("active_col") or mapping.get("is_active") or ""

    if not name_col:
        raise ValueError("Муассаса номи устуни белгиланмаган.")

    rows = []
    for row_number in range(start_row, end_row + 1):
        rows.append({
            "row_number": row_number,
            "name": _get_cell(ws, row_number, name_col),
            "inn": _get_cell(ws, row_number, inn_col),
            "address": _get_cell(ws, row_number, address_col),
            "is_active": _get_cell(ws, row_number, active_col) if active_col else "ҳа",
        })
    return rows


def process_excel_import(uploaded_file, import_type, commit=False, update_existing=False, mapping=None, user=None):
    import_type = import_type or "institutions"
    mapping = mapping or {}

    if import_type not in IMPORT_TYPES:
        return {
            "ok": False,
            "detail": "Бу import тури ҳозирча қўллаб-қувватланмайди.",
            "allowed_import_types": list(IMPORT_TYPES.keys()),
        }

    wb = load_workbook(uploaded_file, data_only=True)

    sheet_name = (mapping.get("sheet_name") or "").strip()
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return {
                "ok": False,
                "detail": f"Excel варақ топилмади: {sheet_name}",
                "sheets": wb.sheetnames,
            }
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    mode = (mapping.get("mode") or "template").strip().lower()

    try:
        if import_type == "institutions":
            if mode in {"manual", "manual_mapping", "mapping"}:
                raw_rows = _read_institution_rows_manual(ws, mapping)
            else:
                raw_rows = _read_institution_rows_template(ws)
        else:
            raw_rows = []
    except Exception as exc:
        return {
            "ok": False,
            "detail": str(exc),
            "sheets": wb.sheetnames,
            "mode": mode,
        }

    results = []
    seen_inn = {}

    for item in raw_rows:
        row_number = item["row_number"]
        inn = _normalize_inn(item.get("inn"))

        if inn and inn in seen_inn:
            results.append(_row_result(
                row_number,
                False,
                "duplicate_in_file",
                "Excel файл ичида ИНН такрорланган.",
                item,
                f"Бу ИНН аввал {seen_inn[inn]}-қаторда учраган.",
            ))
            continue

        if inn:
            seen_inn[inn] = row_number

        results.append(
            _validate_or_save_institution(
                row_number=row_number,
                data=item,
                update_existing=update_existing,
                commit=False,
                user=user,
            )
        )

    validation_summary = _summarize(import_type, results, commit=False)

    if not commit:
        validation_summary["sheets"] = wb.sheetnames
        validation_summary["selected_sheet"] = ws.title
        validation_summary["mode"] = mode
        return validation_summary

    if not validation_summary["ok"]:
        validation_summary["commit"] = False
        validation_summary["rolled_back"] = True
        validation_summary["detail"] = "Хато борлиги учун базага импорт қилинмади."
        return validation_summary

    committed_rows = []
    rolled_back = False

    try:
        with transaction.atomic():
            for item in raw_rows:
                row_number = item["row_number"]

                if not (item.get("name") or item.get("inn")):
                    committed_rows.append(_row_result(row_number, True, "skip", "Бўш қатор ўтказилди.", item))
                    continue

                committed_rows.append(
                    _validate_or_save_institution(
                        row_number=row_number,
                        data=item,
                        update_existing=update_existing,
                        commit=True,
                        user=user,
                    )
                )

            if any(not r.get("ok") for r in committed_rows):
                rolled_back = True
                raise RuntimeError("Import жараёнида хато чиқди.")
    except Exception as exc:
        rolled_back = True
        if not committed_rows:
            committed_rows = results
        summary = _summarize(import_type, committed_rows, commit=False, rolled_back=True)
        summary["detail"] = str(exc)
        return summary

    summary = _summarize(import_type, committed_rows, commit=True, rolled_back=rolled_back)
    summary["sheets"] = wb.sheetnames
    summary["selected_sheet"] = ws.title
    summary["mode"] = mode
    return summary


# --- EXCEL_IMPORT_SAFE_BULK_INSTITUTIONS_V1 ---
# Муассасалар Excel import'ини тезлаштириш:
# preview/validation эски ишчи process_excel_import орқали қолади,
# commit=True ва import_type=institutions бўлганда сақлаш bulk_create/bulk_update билан бажарилади.
# Бошқа import турлари эски логикага fallback қилади.

try:
    _original_process_excel_import_for_safe_bulk = process_excel_import
except NameError:
    _original_process_excel_import_for_safe_bulk = None

if _original_process_excel_import_for_safe_bulk is not None:
    import inspect as _excel_bulk_inspect
    import json as _excel_bulk_json
    from django.db import transaction as _excel_bulk_transaction
    from .models import Institution as _ExcelBulkInstitution

    def _excel_bulk_bound_arguments(uploaded_file, args, kwargs):
        try:
            sig = _excel_bulk_inspect.signature(_original_process_excel_import_for_safe_bulk)
            return sig.bind_partial(uploaded_file, *args, **kwargs)
        except Exception:
            return None

    def _excel_bulk_get_import_type(args, kwargs):
        if "import_type" in kwargs:
            return kwargs.get("import_type")

        if args:
            return args[0]

        return None

    def _excel_bulk_get_commit(uploaded_file, args, kwargs):
        bound = _excel_bulk_bound_arguments(uploaded_file, args, kwargs)

        if bound is not None:
            value = bound.arguments.get("commit", False)
            return bool(value)

        return bool(kwargs.get("commit", False))

    def _excel_bulk_call_original(uploaded_file, args, kwargs, commit):
        kwargs = dict(kwargs)

        try:
            uploaded_file.seek(0)
        except Exception:
            pass

        bound = _excel_bulk_bound_arguments(uploaded_file, args, kwargs)
        if bound is not None:
            bound.arguments["commit"] = commit
            return _original_process_excel_import_for_safe_bulk(*bound.args, **bound.kwargs)

        kwargs["commit"] = commit
        return _original_process_excel_import_for_safe_bulk(uploaded_file, *args, **kwargs)

    def _excel_bulk_rows_from_result(result):
        if not isinstance(result, dict):
            return []

        for key in (
            "rows",
            "preview_rows",
            "results",
            "items",
            "data_rows",
            "preview",
        ):
            value = result.get(key)
            if isinstance(value, list):
                return value

        summary = result.get("summary")
        if isinstance(summary, dict):
            for key in (
                "rows",
                "preview_rows",
                "results",
                "items",
                "data_rows",
                "preview",
            ):
                value = summary.get(key)
                if isinstance(value, list):
                    return value

        return []

    def _excel_bulk_payload_from_row(row):
        if not isinstance(row, dict):
            return None

        for key in (
            "data",
            "payload",
            "object",
            "values",
            "parsed",
            "cleaned_data",
            "detail",
            "message",
            "info",
        ):
            value = row.get(key)

            if isinstance(value, dict):
                return dict(value)

            if isinstance(value, str):
                value = value.strip()
                if value.startswith("{") and value.endswith("}"):
                    try:
                        parsed = _excel_bulk_json.loads(value)
                        if isinstance(parsed, dict):
                            return parsed
                    except Exception:
                        pass

        # Айрим ҳолатларда row'нинг ўзи payload бўлиб келиши мумкин.
        if any(k in row for k in ("name", "inn", "address", "is_active")):
            return dict(row)

        return None

    def _excel_bulk_bool(value, default=True):
        if value is None or value == "":
            return default

        if isinstance(value, bool):
            return value

        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "ha", "ҳа", "xa", "да", "active", "фаол"}:
            return True

        if text in {"0", "false", "no", "yoq", "йўқ", "yo'q", "нет", "inactive", "нофаол"}:
            return False

        return default

    def _excel_bulk_is_ok_row(row):
        if not isinstance(row, dict):
            return False

        if row.get("ok") is False:
            return False

        status = str(row.get("status") or row.get("holat") or row.get("state") or "").strip().lower()
        if status and status not in {"ok", "тўғри", "togri", "to'g'ri", "success", "valid"}:
            return False

        error = row.get("error") or row.get("xato")
        if error and str(error).strip().lower() not in {"ok", "-"}:
            return False

        return True

    def _excel_bulk_commit_institutions_from_preview(preview_result):
        rows = _excel_bulk_rows_from_result(preview_result)

        if not rows:
            return None

        payloads = []
        for row in rows:
            if not _excel_bulk_is_ok_row(row):
                continue

            payload = _excel_bulk_payload_from_row(row)
            if not payload:
                continue

            name = str(payload.get("name") or payload.get("nomi") or "").strip()
            if not name:
                continue

            inn = str(payload.get("inn") or payload.get("ИНН") or payload.get("stir") or "").strip()
            address = str(payload.get("address") or payload.get("manzil") or "").strip()
            is_active = _excel_bulk_bool(payload.get("is_active"), True)

            payloads.append({
                "name": name,
                "inn": inn,
                "address": address,
                "is_active": is_active,
            })

        if not payloads:
            return None

        # Бир Excel файл ичида такрор келган INN/name бўлса, охирги қатор устун туради.
        normalized = {}
        for item in payloads:
            key = ("inn", item["inn"]) if item["inn"] else ("name", item["name"].casefold())
            normalized[key] = item

        payloads = list(normalized.values())

        existing_by_inn = {
            inst.inn: inst
            for inst in _ExcelBulkInstitution.objects.filter(
                inn__in=[item["inn"] for item in payloads if item["inn"]]
            )
        }

        existing_by_name = {
            inst.name.casefold(): inst
            for inst in _ExcelBulkInstitution.objects.filter(
                name__in=[item["name"] for item in payloads if not item["inn"]]
            )
        }

        to_create = []
        to_update = []

        for item in payloads:
            existing = None

            if item["inn"]:
                existing = existing_by_inn.get(item["inn"])

            if existing is None and not item["inn"]:
                existing = existing_by_name.get(item["name"].casefold())

            if existing is None:
                to_create.append(_ExcelBulkInstitution(**item))
            else:
                existing.name = item["name"]
                existing.inn = item["inn"]
                existing.address = item["address"]
                existing.is_active = item["is_active"]
                to_update.append(existing)

        with _excel_bulk_transaction.atomic():
            if to_create:
                _ExcelBulkInstitution.objects.bulk_create(to_create, batch_size=1000)

            if to_update:
                _ExcelBulkInstitution.objects.bulk_update(
                    to_update,
                    ["name", "inn", "address", "is_active"],
                    batch_size=1000,
                )

        result = dict(preview_result)
        summary = dict(result.get("summary") or {})

        summary["created"] = len(to_create)
        summary["updated"] = len(to_update)
        summary["errors"] = int(summary.get("errors") or summary.get("error") or 0)
        summary["ok"] = len(payloads)

        result["summary"] = summary
        result["created"] = len(to_create)
        result["updated"] = len(to_update)
        result["errors"] = summary["errors"]
        result["ok"] = len(payloads)
        result["committed"] = True
        result["bulk_optimized"] = True
        result["bulk_model"] = "Institution"

        return result

    def process_excel_import(uploaded_file, *args, **kwargs):
        import_type = str(_excel_bulk_get_import_type(args, kwargs) or "institutions").strip()
        import_type_norm = import_type.lower().replace("-", "_")
        commit = _excel_bulk_get_commit(uploaded_file, args, kwargs)

        if not commit or import_type_norm not in {"institutions", "institution", "muassasalar", "муссасалар", "муассасалар"}:
            return _original_process_excel_import_for_safe_bulk(uploaded_file, *args, **kwargs)

        preview_result = _excel_bulk_call_original(uploaded_file, args, kwargs, commit=False)

        try:
            fast_result = _excel_bulk_commit_institutions_from_preview(preview_result)
        except Exception:
            fast_result = None

        if fast_result is not None:
            return fast_result

        # Preview shape мос келмаса, эски ишчи логика бузилмасин.
        return _excel_bulk_call_original(uploaded_file, args, kwargs, commit=True)
# --- /EXCEL_IMPORT_SAFE_BULK_INSTITUTIONS_V1 ---

